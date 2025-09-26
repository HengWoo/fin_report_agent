"""
Financial Symbol Navigator for LSP-like Account Traversal

This module provides LSP-inspired navigation through financial account structures,
enabling intelligent exploration of financial data similar to how Serena navigates code.

Features:
- Hierarchical account navigation (like LSP symbols)
- Account relationship discovery
- Context-aware financial data retrieval
- Pattern matching for account search
"""

import logging
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook


@dataclass
class AccountSymbol:
    """Financial account symbol (analogous to LSP symbol)."""

    name: str
    name_path: str  # Full path like "资产/流动资产/现金"
    account_type: str  # asset, liability, revenue, expense, etc.
    level: int  # Hierarchy level (0 = root)
    line_number: Optional[int] = None
    column: Optional[str] = None
    parent_path: Optional[str] = None
    children: List[str] = field(default_factory=list)
    values: Dict[str, Any] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def is_leaf(self) -> bool:
        """Check if this is a leaf account (no children)."""
        return len(self.children) == 0

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "name": self.name,
            "name_path": self.name_path,
            "account_type": self.account_type,
            "level": self.level,
            "line_number": self.line_number,
            "column": self.column,
            "parent_path": self.parent_path,
            "children": self.children,
            "is_leaf": self.is_leaf(),
            "values": self.values,
            "metadata": self.metadata,
        }


class FinancialSymbolNavigator:
    """Navigator for hierarchical financial account structures."""

    def __init__(self):
        """Initialize the navigator."""
        self.logger = logging.getLogger("financial_navigator")

        # Cached account structures by file
        self.account_trees: Dict[str, Dict[str, AccountSymbol]] = {}

        # Common Chinese account patterns
        self.account_patterns = {
            "asset": ["资产", "asset"],
            "liability": ["负债", "liability"],
            "revenue": ["收入", "revenue", "营业收入"],
            "expense": ["支出", "费用", "expense", "成本"],
            "equity": ["所有者权益", "equity"],
            "profit": ["利润", "profit", "净利润"],
        }

    def parse_excel_structure(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> Dict[str, AccountSymbol]:
        """Parse Excel file structure into account symbols."""
        if file_path in self.account_trees:
            return self.account_trees[file_path]

        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active

            accounts: Dict[str, AccountSymbol] = {}
            parent_stack: List[tuple[int, str]] = []  # (level, name_path)

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if not row or not row[0]:
                    continue

                account_name = str(row[0]).strip()

                # Skip headers and empty rows
                if not account_name or account_name in [
                    "项目",
                    "账户",
                    "科目",
                    "Account",
                ]:
                    continue

                # Determine hierarchy level (by indentation or numbering)
                level = self._detect_hierarchy_level(account_name, row_idx)

                # Build name path
                while parent_stack and parent_stack[-1][0] >= level:
                    parent_stack.pop()

                parent_path = parent_stack[-1][1] if parent_stack else None
                name_path = (
                    f"{parent_path}/{account_name}" if parent_path else account_name
                )

                # Extract values from columns
                values = {}
                for col_idx, value in enumerate(row[1:], start=1):
                    if value is not None and isinstance(value, (int, float)):
                        values[f"col_{col_idx}"] = value

                # Detect account type
                account_type = self._detect_account_type(account_name, name_path)

                # Create symbol
                symbol = AccountSymbol(
                    name=account_name,
                    name_path=name_path,
                    account_type=account_type,
                    level=level,
                    line_number=row_idx,
                    parent_path=parent_path,
                    values=values,
                    metadata={"original_row": row_idx},
                )

                accounts[name_path] = symbol

                # Update parent's children
                if parent_path and parent_path in accounts:
                    accounts[parent_path].children.append(name_path)

                # Update stack
                parent_stack.append((level, name_path))

            self.account_trees[file_path] = accounts
            self.logger.info(f"Parsed {len(accounts)} account symbols from {file_path}")

            return accounts

        except Exception as e:
            self.logger.error(f"Failed to parse Excel structure: {str(e)}")
            return {}

    def _detect_hierarchy_level(self, account_name: str, row_idx: int) -> int:
        """Detect hierarchy level from account name or numbering."""
        # Common Chinese numbering: 一、二、三、 or (一)、(二)、
        if account_name.startswith(
            (
                "一、",
                "二、",
                "三、",
                "四、",
                "五、",
                "六、",
                "七、",
                "八、",
                "九、",
                "十、",
            )
        ):
            return 0

        # Sub-levels: (一)、(二)、 or 1.、2.、
        if account_name.startswith("(") or account_name[0].isdigit():
            if "(" in account_name[:5]:
                return 1
            else:
                return 2

        # Indentation-based (count leading spaces)
        leading_spaces = len(account_name) - len(account_name.lstrip())
        if leading_spaces > 0:
            return (leading_spaces // 2) + 1

        # Default: top level
        return 0

    def _detect_account_type(self, account_name: str, name_path: str) -> str:
        """Detect account type from name and path."""
        for acc_type, patterns in self.account_patterns.items():
            for pattern in patterns:
                if pattern in account_name.lower() or pattern in name_path.lower():
                    return acc_type

        return "unknown"

    def find_account(
        self,
        file_path: str,
        name_pattern: str,
        exact_match: bool = False,
        account_type: Optional[str] = None,
    ) -> List[AccountSymbol]:
        """Find accounts matching name pattern (like LSP find_symbol)."""
        accounts = self.parse_excel_structure(file_path)

        results = []
        for symbol in accounts.values():
            # Name matching
            if exact_match:
                name_match = (
                    symbol.name == name_pattern or symbol.name_path == name_pattern
                )
            else:
                name_match = (
                    name_pattern.lower() in symbol.name.lower()
                    or name_pattern.lower() in symbol.name_path.lower()
                )

            # Type filtering
            type_match = True
            if account_type:
                type_match = symbol.account_type == account_type

            if name_match and type_match:
                results.append(symbol)

        return results

    def get_financial_overview(
        self, file_path: str, max_depth: int = 2
    ) -> List[AccountSymbol]:
        """Get high-level financial structure overview (like LSP get_symbols_overview)."""
        accounts = self.parse_excel_structure(file_path)

        overview = [symbol for symbol in accounts.values() if symbol.level <= max_depth]

        return sorted(overview, key=lambda x: (x.level, x.line_number or 0))

    def find_referencing_accounts(
        self, file_path: str, account_name_path: str
    ) -> List[Dict[str, Any]]:
        """Find accounts that reference/depend on the given account (like LSP find_referencing_symbols)."""
        accounts = self.parse_excel_structure(file_path)

        if account_name_path not in accounts:
            return []

        target = accounts[account_name_path]
        references = []

        # Find children (direct references)
        for child_path in target.children:
            if child_path in accounts:
                child = accounts[child_path]
                references.append(
                    {
                        "type": "child",
                        "account": child.to_dict(),
                        "relationship": "part_of",
                    }
                )

        # Find parent (referenced by)
        if target.parent_path and target.parent_path in accounts:
            parent = accounts[target.parent_path]
            references.append(
                {
                    "type": "parent",
                    "account": parent.to_dict(),
                    "relationship": "contains",
                }
            )

        return references

    def get_account_context(
        self, file_path: str, account_name_path: str, depth: int = 1
    ) -> Dict[str, Any]:
        """Get account with surrounding context (like reading with context lines)."""
        accounts = self.parse_excel_structure(file_path)

        if account_name_path not in accounts:
            return {"error": "Account not found"}

        target = accounts[account_name_path]

        context = {
            "account": target.to_dict(),
            "children": [],
            "siblings": [],
            "ancestors": [],
        }

        # Get children up to depth
        if depth >= 1:
            context["children"] = [
                accounts[child].to_dict()
                for child in target.children
                if child in accounts
            ]

        # Get siblings (same parent)
        if target.parent_path:
            parent = accounts.get(target.parent_path)
            if parent:
                context["siblings"] = [
                    accounts[sibling].to_dict()
                    for sibling in parent.children
                    if sibling != account_name_path and sibling in accounts
                ]

        # Get ancestors (parent chain)
        current_path = target.parent_path
        while current_path and current_path in accounts:
            ancestor = accounts[current_path]
            context["ancestors"].append(ancestor.to_dict())
            current_path = ancestor.parent_path

        return context

    def get_leaf_accounts(
        self, file_path: str, account_type: Optional[str] = None
    ) -> List[AccountSymbol]:
        """Get all leaf accounts (accounts with no children) for safe calculations."""
        accounts = self.parse_excel_structure(file_path)

        leaf_accounts = [symbol for symbol in accounts.values() if symbol.is_leaf()]

        if account_type:
            leaf_accounts = [a for a in leaf_accounts if a.account_type == account_type]

        return leaf_accounts

    def get_account_hierarchy(
        self, file_path: str, root_pattern: Optional[str] = None
    ) -> Dict[str, Any]:
        """Get hierarchical account tree structure."""
        accounts = self.parse_excel_structure(file_path)

        # Find root accounts
        if root_pattern:
            roots = [
                a
                for a in accounts.values()
                if root_pattern.lower() in a.name.lower() and a.level == 0
            ]
        else:
            roots = [a for a in accounts.values() if a.level == 0]

        def build_tree(account: AccountSymbol) -> Dict[str, Any]:
            """Recursively build account tree."""
            node = account.to_dict()
            node["children"] = [
                build_tree(accounts[child])
                for child in account.children
                if child in accounts
            ]
            return node

        hierarchy = {
            "file_path": file_path,
            "total_accounts": len(accounts),
            "root_accounts": [build_tree(root) for root in roots],
            "account_types": self._get_type_summary(accounts),
        }

        return hierarchy

    def _get_type_summary(self, accounts: Dict[str, AccountSymbol]) -> Dict[str, int]:
        """Get summary of account types."""
        type_counts: Dict[str, int] = {}

        for symbol in accounts.values():
            acc_type = symbol.account_type
            type_counts[acc_type] = type_counts.get(acc_type, 0) + 1

        return type_counts

    def search_accounts_by_value(
        self,
        file_path: str,
        min_value: Optional[float] = None,
        max_value: Optional[float] = None,
        column: Optional[str] = None,
    ) -> List[AccountSymbol]:
        """Search accounts by value range."""
        accounts = self.parse_excel_structure(file_path)

        results = []
        for symbol in accounts.values():
            if not symbol.values:
                continue

            values_to_check = (
                [symbol.values[column]]
                if column and column in symbol.values
                else symbol.values.values()
            )

            for value in values_to_check:
                if isinstance(value, (int, float)):
                    if min_value is not None and value < min_value:
                        continue
                    if max_value is not None and value > max_value:
                        continue

                    results.append(symbol)
                    break

        return results

    def get_account_path_chain(
        self, file_path: str, account_name_path: str
    ) -> List[AccountSymbol]:
        """Get full path chain from root to account."""
        accounts = self.parse_excel_structure(file_path)

        if account_name_path not in accounts:
            return []

        chain = []
        current = accounts[account_name_path]

        while current:
            chain.insert(0, current)
            if current.parent_path and current.parent_path in accounts:
                current = accounts[current.parent_path]
            else:
                break

        return chain


# Global instance
financial_navigator = FinancialSymbolNavigator()
